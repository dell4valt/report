import os
import sys
import unittest
from pathlib import Path
from unittest.mock import patch, MagicMock
import tempfile
import pandas as pd
from openpyxl import Workbook

# Получаем абсолютный путь к основной папке проекта
# и добавляем его в переменную среды PYTHONPATH
project_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(project_root))

from report.utils import *


class TestGetXlsSheetQuantity(unittest.TestCase):
    """Тесты для функции get_xls_sheet_quantity"""

    def setUp(self):
        """Подготовка тестов"""
        # Создаем временный тестовый файл Excel
        self.temp_dir = tempfile.TemporaryDirectory()
        self.test_file_path = os.path.join(self.temp_dir.name, "test_file.xlsx")

        # Создаем новый Excel файл с двумя листами
        wb = Workbook()
        wb.create_sheet("Sheet1")
        wb.create_sheet("Sheet2")
        wb.save(self.test_file_path)

    def tearDown(self):
        """Очистка после тестов"""
        self.temp_dir.cleanup()

    def test_get_sheets_count(self):
        """Тест подсчета количества листов"""
        sheet_count = get_xls_sheet_quantity(self.test_file_path)
        self.assertEqual(sheet_count, 3)

    def test_file_not_found(self):
        """Тест обработки отсутствующего файла"""
        non_existent_file = "not_existing_file.xlsx"
        with self.assertRaises(FileNotFoundError):
            get_xls_sheet_quantity(non_existent_file)

    @patch("report.utils.load_workbook")
    def test_with_mock(self, mock_load_workbook):
        """Тест с имитацией загрузки файла"""
        # Настраиваем мок для имитации файла с 3 листами
        mock_workbook = MagicMock()
        mock_workbook.sheetnames = ["Sheet1", "Sheet2", "Sheet3"]
        mock_load_workbook.return_value = mock_workbook

        # Вызываем функцию с любым путем, т.к. мы используем мок
        sheet_count = get_xls_sheet_quantity("any_path.xlsx")
        self.assertEqual(sheet_count, 3)


class TestInsertRowNumbersInDF(unittest.TestCase):
    """Тесты для функции insert_row_numbers_in_df"""

    def test_empty_dataframe(self):
        """Тест с пустым DataFrame"""
        df = pd.DataFrame()
        result = insert_row_numbers_in_df(df)
        self.assertEqual(len(result), 0)
        self.assertTrue("№" in result.columns)

    def test_with_data(self):
        """Тест с наполненным DataFrame"""
        # Создаём тестовый DataFrame
        df = pd.DataFrame({"A": [1, 2, 3], "B": ["a", "b", "c"]})

        # Проверяем результат
        result = insert_row_numbers_in_df(df)
        self.assertEqual(len(result), 3)
        self.assertTrue("№" in result.columns)
        self.assertEqual(list(result["№"]), [1, 2, 3])

        # Проверяем, что исходные данные сохранены
        self.assertTrue("A" in result.columns)
        self.assertTrue("B" in result.columns)
        self.assertEqual(list(result["A"]), [1, 2, 3])
        self.assertEqual(list(result["B"]), ["a", "b", "c"])

    def test_custom_index_name(self):
        """Тест с настраиваемым именем индекса"""
        df = pd.DataFrame({"X": [10, 20]})
        custom_name = "Index"
        result = insert_row_numbers_in_df(df, name=custom_name)

        self.assertTrue(custom_name in result.columns)
        self.assertEqual(list(result[custom_name]), [1, 2])


if __name__ == "__main__":
    unittest.main()
