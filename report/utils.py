from openpyxl import load_workbook
import pandas as pd


def get_xls_sheet_quantity(file_path) -> int:
    """Функция возвращает количество листов в указанном xls файле.

    Args:
        file_path (str): Путь к xls файлу

    Returns:
        int: Количество листов в файле
    """
    try:
        data_file = load_workbook(file_path)
    except FileNotFoundError:
        raise FileNotFoundError(
            "Ошибка определения количества листов в Excel файле! "
            f"Файл {file_path} не найден. Программа будет завершена."
        )

    return len(data_file.sheetnames)


def insert_row_numbers_in_df(df: pd.DataFrame, name: str = "№") -> pd.DataFrame:
    """Функция вставляет колонку индекса строк в DataFrame.

    Args:
        df (pd.DataFrame): DataFrame с данными
        name (str): Название колонки индекса

    Returns:
        pd.DataFrame: DataFrame с вставленным индексом
    """
    # Проверяем, пустой ли DataFrame
    if df.empty:
        # Создаем новый DataFrame с одним столбцом для номеров строк
        return pd.DataFrame(columns=[name])

    row_numbers = pd.Series(range(1, len(df) + 1), name=name)
    df = pd.concat([row_numbers, df.reset_index(drop=True)], axis=1)
    return df
